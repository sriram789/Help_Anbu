# importing openpyxl module 
import openpyxl 
  
# Give the location of the file 
path = "/Users/sriramn/Desktop/Anbu_Project/Anbu_IP.xlsx"

#Creating a set for Storing Description Elements
Description_set={}
# workbook object is created 
read_wb_obj = openpyxl.load_workbook(path) 
  
read_sheet_obj = read_wb_obj.active 

m_row = read_sheet_obj.max_row 
  
# Loop will print all values 
# of first column  
for i in range(3, m_row + 1): 
    cell_obj = read_sheet_obj.cell(row = i, column = 2) 
    if(cell_obj.value != 'None'):
    	#print(cell_obj.value)
    	if(isinstance(Description_set.get(cell_obj.value),int)):
    		count=Description_set.get(cell_obj.value)
    		Description_set[cell_obj.value]=int(count+1)
    	else:
    		Description_set[cell_obj.value]=int(0)

#writing into another excel file
wb = openpyxl.Workbook() 
  
# Get workbook active sheet   
# from the active attribute 
sheet = wb.active 
  
# Cell objects also have row, column 
# and coordinate attributes that provide 
# location information for the cell. 
  
# Note: The first row or column integer 
# is 1, not 0. Cell object is created by 
# using sheet object's cell() method.
iterator = 1 
for Description_name, count in Description_set.items():

	print(Description_name+"    "+str(count)+"\n")
	Description_column = sheet.cell(row = iterator, column = 1) 
	Count_column = sheet.cell(row = iterator, column = 2)
	iterator+=1
	Description_column.value = Description_name
	Count_column.value = count

wb.save("/Users/sriramn/Desktop/Anbu_Project/Anbu_OP.xlsx")

